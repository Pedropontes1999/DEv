import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import threading
import time
import math
import cv2
import mediapipe as mp


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def detectar_encoding(caminho):
    for enc in ("windows-1252", "utf-8", "latin-1"):
        try:
            with open(caminho, encoding=enc) as f:
                f.read(1024)
            return enc
        except UnicodeDecodeError:
            pass
    return "windows-1252"


def ler_csv(caminho):
    enc = detectar_encoding(caminho)
    # Tenta leitura normal primeiro
    try:
        return pd.read_csv(caminho, sep=";", encoding=enc, dtype=str,
                           on_bad_lines="skip", engine="python")
    except Exception:
        pass
    # Fallback: lê linha a linha ignorando linhas problemáticas
    try:
        return pd.read_csv(caminho, sep=";", encoding=enc, dtype=str,
                           on_bad_lines="skip", engine="python",
                           quoting=3)  # QUOTE_NONE
    except Exception:
        pass
    # Último recurso: lê como texto e reconstrói
    linhas = []
    with open(caminho, encoding=enc, errors="replace") as f:
        raw = f.readlines()
    if not raw:
        raise ValueError("Arquivo vazio.")
    header = raw[0].strip().split(";")
    ncols = len(header)
    for linha in raw[1:]:
        partes = linha.strip().split(";")
        if len(partes) >= ncols:
            linhas.append(partes[:ncols])
        elif len(partes) > 0:
            # Completa com vazio
            partes += [""] * (ncols - len(partes))
            linhas.append(partes)
    return pd.DataFrame(linhas, columns=header)


def parse_estoque(valor):
    result = {}
    if not isinstance(valor, str) or not valor.strip():
        return result
    for parte in valor.split("|"):
        parte = parte.strip()
        if "," in parte:
            try:
                num, qtd = parte.split(",", 1)
                result[num.strip()] = int(qtd.strip())
            except ValueError:
                pass
    return result


def formatar_estoque(dicionario):
    return "|".join(f"{n},{q}" for n, q in sorted(dicionario.items(), key=lambda x: float(x[0])))


def calcular_diferenca(est_novo, est_antigo):
    numeros = sorted(set(est_novo) | set(est_antigo), key=lambda x: float(x))
    partes = []
    for n in numeros:
        q_novo = est_novo.get(n, 0)
        q_antigo = est_antigo.get(n, 0)
        diff = q_novo - q_antigo
        sinal = "+" if diff > 0 else ""
        partes.append(f"{n},{sinal}{diff}")
    return "|".join(partes)


def total_estoque(dicionario):
    return sum(dicionario.values())


# ─────────────────────────────────────────────
#  FUNÇÃO 1 – COMPARAR PLANILHAS
# ─────────────────────────────────────────────

def comparar_planilhas(path_novo, path_antigo, path_saida):
    df_novo = ler_csv(path_novo)
    df_antigo = ler_csv(path_antigo)
    df_novo.columns = [c.strip().upper() for c in df_novo.columns]
    df_antigo.columns = [c.strip().upper() for c in df_antigo.columns]

    col_ref = None
    for candidato in ("REFERÊNCIA", "REFERENCIA", "REF", "CODIGO", "ID"):
        if candidato in df_novo.columns:
            col_ref = candidato
            break
    if col_ref is None:
        col_ref = df_novo.columns[0]

    col_est = "ESTOQUE"
    if col_est not in df_novo.columns:
        raise ValueError(f"Coluna ESTOQUE não encontrada. Colunas: {list(df_novo.columns)}")

    merged = df_novo.merge(
        df_antigo[[col_ref, col_est]],
        on=col_ref,
        how="outer",
        suffixes=("_NOVO", "_ANTIGO")
    )

    linhas = []
    for _, row in merged.iterrows():
        est_novo = parse_estoque(row.get(f"{col_est}_NOVO") or row.get(col_est, ""))
        est_antigo = parse_estoque(row.get(f"{col_est}_ANTIGO", ""))
        total_novo = total_estoque(est_novo)
        total_antigo = total_estoque(est_antigo)
        variacao_total = total_novo - total_antigo

        if variacao_total > 0:
            status = "Aumentou"
        elif variacao_total < 0:
            status = "Diminuiu"
        elif total_novo == 0:
            status = "Zerado"
        else:
            status = "Igual"

        linha = row.to_dict()
        for c in [f"{col_est}_NOVO", f"{col_est}_ANTIGO", col_est]:
            if c in linha:
                del linha[c]

        # Números que saíram do zero (liberados para venda)
        def sort_nums(lst):
            try:
                return sorted(lst, key=lambda x: float(x))
            except Exception:
                return sorted(lst)

        liberados = [n for n in set(list(est_novo.keys()) + list(est_antigo.keys()))
                     if est_antigo.get(n, 0) == 0 and est_novo.get(n, 0) > 0]
        esgotados = [n for n in set(list(est_novo.keys()) + list(est_antigo.keys()))
                     if est_antigo.get(n, 0) > 0 and est_novo.get(n, 0) == 0]

        linha["ESTOQUE_ATUAL"] = formatar_estoque(est_novo)
        linha["ESTOQUE_ANTERIOR"] = formatar_estoque(est_antigo)
        linha["DIFERENCA_POR_NUMERO"] = calcular_diferenca(est_novo, est_antigo)
        linha["TOTAL_ATUAL"] = total_novo
        linha["TOTAL_ANTERIOR"] = total_antigo
        linha["VARIACAO_TOTAL"] = variacao_total
        linha["STATUS"] = status
        linha["NUMEROS_LIBERADOS"] = "|".join(sort_nums(liberados)) if liberados else ""
        linha["NUMEROS_ESGOTADOS"] = "|".join(sort_nums(esgotados)) if esgotados else ""
        linhas.append(linha)

    df_saida = pd.DataFrame(linhas)
    colunas_est = ["ESTOQUE_ATUAL", "ESTOQUE_ANTERIOR", "DIFERENCA_POR_NUMERO",
                   "TOTAL_ATUAL", "TOTAL_ANTERIOR", "VARIACAO_TOTAL", "STATUS",
                   "NUMEROS_LIBERADOS", "NUMEROS_ESGOTADOS"]
    outras = [c for c in df_saida.columns if c not in colunas_est]
    df_saida = df_saida[outras + colunas_est]

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df_saida.to_excel(path_saida, index=False, engine="openpyxl")
    wb = load_workbook(path_saida)
    ws = wb.active

    cor_map = {
        "Aumentou": "C6EFCE",
        "Diminuiu": "FFCCCC",
        "Zerado":   "FFEB9C",
        "Igual":    "FFFFFF",
    }

    header_fill = PatternFill("solid", fgColor="0F2D4A")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Mapeia índices das colunas importantes
    col_idx = {}
    for i, cell in enumerate(ws[1], 1):
        col_idx[cell.value] = i

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    FILL_STATUS   = {k: PatternFill("solid", fgColor=v) for k, v in cor_map.items()}
    FILL_AMARELO  = PatternFill("solid", fgColor="FFD700")   # liberado para venda
    FILL_VERMELHO = PatternFill("solid", fgColor="FF9999")   # esgotado
    FILL_BRANCO   = PatternFill("solid", fgColor="FFFFFF")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_val    = str(row[col_idx.get("STATUS", 1) - 1].value or "") if "STATUS" in col_idx else ""
        tem_liberado  = bool(str(row[col_idx.get("NUMEROS_LIBERADOS", 1) - 1].value or "").strip()) if "NUMEROS_LIBERADOS" in col_idx else False
        tem_esgotado  = bool(str(row[col_idx.get("NUMEROS_ESGOTADOS", 1) - 1].value or "").strip()) if "NUMEROS_ESGOTADOS" in col_idx else False

        for cell in row:
            cell.border = border
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")

            col_name = ws.cell(row=1, column=cell.col_idx).value

            if col_name == "STATUS":
                cell.fill = FILL_STATUS.get(status_val, FILL_BRANCO)

            elif col_name in ("ESTOQUE_ATUAL", "ESTOQUE_ANTERIOR"):
                # Prioridade: liberado > esgotado > padrão branco
                if tem_liberado:
                    cell.fill = FILL_AMARELO
                    cell.font = Font(name="Calibri", size=9, bold=True, color="7D4E00")
                elif tem_esgotado:
                    cell.fill = FILL_VERMELHO
                    cell.font = Font(name="Calibri", size=9, bold=True, color="8B0000")

            elif col_name == "NUMEROS_LIBERADOS" and tem_liberado:
                cell.fill = FILL_AMARELO
                cell.font = Font(name="Calibri", size=9, bold=True, color="7D4E00")

            elif col_name == "NUMEROS_ESGOTADOS" and tem_esgotado:
                cell.fill = FILL_VERMELHO
                cell.font = Font(name="Calibri", size=9, bold=True, color="8B0000")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path_saida)


# ─────────────────────────────────────────────
#  FUNÇÃO 2 – CSV PARA EXCEL
# ─────────────────────────────────────────────

def csv_para_excel(path_csv, path_saida):
    df = ler_csv(path_csv)
    df.columns = [c.strip() for c in df.columns]

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    df.to_excel(path_saida, index=False, engine="openpyxl")
    wb = load_workbook(path_saida)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="0F2D4A")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path_saida)


# ─────────────────────────────────────────────
#  GESTOS — MediaPipe
# ─────────────────────────────────────────────

PINCH_THRESH      = 0.055   # distância normalizada polegar↔indicador
GESTURE_COOLDOWN  = 2.0     # segundos entre ativações


def _is_finger_extended(lm, tip, pip):
    """Retorna True se o dedo está estendido (ponta acima da articulação PIP)."""
    return lm.landmark[tip].y < lm.landmark[pip].y


def detectar_pincha(lm):
    """Pinça simples: polegar + indicador juntos, outros dedos dobrados."""
    t = lm.landmark[4]
    i = lm.landmark[8]
    pinch = math.hypot(t.x - i.x, t.y - i.y) < PINCH_THRESH
    # Ao menos 2 dos outros dedos dobrados
    medio_ext  = _is_finger_extended(lm, 12, 10)
    anelar_ext = _is_finger_extended(lm, 16, 14)
    mindinho_ext = _is_finger_extended(lm, 20, 18)
    n_estendidos = sum([medio_ext, anelar_ext, mindinho_ext])
    return pinch and n_estendidos <= 1


def detectar_ok(lm):
    """Sinal de OK: polegar + indicador formando círculo, outros 3 dedos estendidos."""
    t = lm.landmark[4]
    i = lm.landmark[8]
    circulo = math.hypot(t.x - i.x, t.y - i.y) < PINCH_THRESH
    medio_ext    = _is_finger_extended(lm, 12, 10)
    anelar_ext   = _is_finger_extended(lm, 16, 14)
    mindinho_ext = _is_finger_extended(lm, 20, 18)
    return circulo and medio_ext and anelar_ext and mindinho_ext


class GestureController:
    """Roda em thread separada, detecta gestos e dispara callbacks no tkinter."""

    def __init__(self, app):
        self.app     = app
        self.running = False
        self._thread = None

    def start(self):
        self.running = True
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def stop(self):
        self.running = False

    def _loop(self):
        mp_hands = mp.solutions.hands
        hands = mp_hands.Hands(
            max_num_hands=1,
            min_detection_confidence=0.70,
            min_tracking_confidence=0.60,
        )
        cap = cv2.VideoCapture(0)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  320)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 240)

        last_trigger = 0.0
        prev_pinch   = False
        prev_ok      = False

        while self.running:
            ret, frame = cap.read()
            if not ret:
                break

            frame = cv2.flip(frame, 1)
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            rgb.flags.writeable = False
            res = hands.process(rgb)
            rgb.flags.writeable = True

            pinch_now = ok_now = False

            if res.multi_hand_landmarks:
                lm = res.multi_hand_landmarks[0]
                ok_now    = detectar_ok(lm)
                pinch_now = detectar_pincha(lm) and not ok_now

                # Desenhar landmarks
                mp.solutions.drawing_utils.draw_landmarks(
                    frame, lm, mp_hands.HAND_CONNECTIONS)

            now = time.time()

            # ── Pinça → selecionar arquivos ──
            if pinch_now and not prev_pinch:
                if now - last_trigger > GESTURE_COOLDOWN:
                    last_trigger = now
                    self.app.after(0, self.app.gesto_selecionar)

            # ── OK → executar macro ──
            if ok_now and not prev_ok:
                if now - last_trigger > GESTURE_COOLDOWN:
                    last_trigger = now
                    self.app.after(0, self.app.gesto_executar)

            prev_pinch = pinch_now
            prev_ok    = ok_now

            # ── Overlay visual ──
            if ok_now:
                gesto_txt = "OK — EXECUTAR"
                cor = (0, 220, 120)
            elif pinch_now:
                gesto_txt = "PINCA — SELECIONAR"
                cor = (0, 200, 255)
            else:
                gesto_txt = "Aguardando gesto..."
                cor = (140, 140, 140)

            cv2.rectangle(frame, (0, 0), (320, 30), (15, 25, 40), -1)
            cv2.putText(frame, gesto_txt, (8, 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, cor, 1, cv2.LINE_AA)

            cv2.imshow("Gestos — Estoque", frame)
            if cv2.waitKey(1) == 27:
                self.running = False
                break

        cap.release()
        hands.close()
        cv2.destroyAllWindows()


# ─────────────────────────────────────────────
#  CORES
# ─────────────────────────────────────────────

BG       = "#0F1923"
SURFACE  = "#162230"
SURFACE2 = "#1C2D3E"
BORDER   = "#243548"
ACCENT   = "#00C2FF"
ACCENT2  = "#009ECC"
SUCCESS  = "#00E5A0"
WARNING  = "#FFB800"
DANGER   = "#FF4D6A"
TEXT     = "#E8F0FA"
TEXT_DIM = "#7A95B0"
TEXT_MUT = "#3D5570"
INP_BG   = "#0A1520"
BTN_FG   = "#0A1520"


# ─────────────────────────────────────────────
#  BUTTON HELPER – usa tk.Button nativo
# ─────────────────────────────────────────────

def make_button(parent, text, command, primary=True, width=18):
    """Cria um botão estilizado usando tk.Button nativo (100% compatível Windows)."""
    if primary:
        bg, fg, ab = ACCENT, BTN_FG, ACCENT2
    else:
        bg, fg, ab = SURFACE2, TEXT, BORDER

    btn = tk.Button(
        parent,
        text=text,
        command=command,
        bg=bg,
        fg=fg,
        activebackground=ab,
        activeforeground=fg,
        relief="flat",
        bd=0,
        font=("Segoe UI", 9, "bold"),
        cursor="hand2",
        width=width,
        pady=8,
    )
    # Hover manual
    btn.bind("<Enter>", lambda e: btn.config(bg=ab))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


# ─────────────────────────────────────────────
#  WIDGET: FileRow
# ─────────────────────────────────────────────

class FileRow(tk.Frame):
    def __init__(self, parent, label, var, save_mode=False, **kw):
        super().__init__(parent, bg=BG, **kw)
        self.var = var

        tk.Label(self, text=label, font=("Segoe UI", 8, "bold"),
                 bg=BG, fg=TEXT_DIM).pack(anchor="w", pady=(12, 3))

        row = tk.Frame(self, bg=BG)
        row.pack(fill="x")

        # Borda do entry
        wrap = tk.Frame(row, bg=BORDER, padx=1, pady=1)
        wrap.pack(side="left", fill="x", expand=True, padx=(0, 8))
        inner = tk.Frame(wrap, bg=INP_BG)
        inner.pack(fill="both", expand=True)
        self.entry = tk.Entry(inner, textvariable=var, bg=INP_BG, fg=TEXT,
                              insertbackground=ACCENT, relief="flat",
                              font=("Segoe UI", 9), bd=0)
        self.entry.pack(fill="x", padx=10, pady=7)
        self.entry.bind("<FocusIn>",  lambda e: wrap.config(bg=ACCENT))
        self.entry.bind("<FocusOut>", lambda e: wrap.config(bg=BORDER))

        if save_mode:
            btn = make_button(row, "Salvar em...", self._save, primary=False, width=14)
        else:
            btn = make_button(row, "Selecionar",  self._open, primary=False, width=14)
        btn.pack(side="left")

    def _open(self):
        p = filedialog.askopenfilename(
            filetypes=[("CSV / Excel", "*.csv *.xlsx *.xls"), ("Todos", "*.*")])
        if p:
            self.var.set(p)

    def _save(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if p:
            self.var.set(p)


# ─────────────────────────────────────────────
#  WIDGET: LogBox
# ─────────────────────────────────────────────

class LogBox(tk.Frame):
    def __init__(self, parent, lines=5, **kw):
        super().__init__(parent, bg=BORDER, padx=1, pady=1, **kw)
        inner = tk.Frame(self, bg=INP_BG)
        inner.pack(fill="both", expand=True)
        self.txt = tk.Text(inner, height=lines, bg=INP_BG, fg=SUCCESS,
                           font=("Consolas", 8), relief="flat", bd=0,
                           state="disabled", wrap="word", padx=12, pady=8,
                           cursor="arrow")
        self.txt.pack(fill="both", expand=True)
        self.txt.tag_config("err",  foreground=DANGER)
        self.txt.tag_config("info", foreground=ACCENT)
        self.txt.tag_config("dim",  foreground=TEXT_MUT)
        self.txt.tag_config("ok",   foreground=SUCCESS)

    def log(self, msg, tag="ok"):
        self.txt.config(state="normal")
        self.txt.insert("end", msg + "\n", tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def clear(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.config(state="disabled")


# ─────────────────────────────────────────────
#  APP PRINCIPAL
# ─────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gestao de Estoque de Calcados")
        self.geometry("700x610")
        self.resizable(False, False)
        self.configure(bg=BG)
        self._active_tab = "comparar"

        self._build_header()
        self._build_gesture_bar()
        self._build_tabs()

        self._gesture = GestureController(self)
        self._gesture.start()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        self._gesture.stop()
        self.destroy()

    # ── BARRA DE GESTOS ─────────────────────

    def _build_gesture_bar(self):
        bar = tk.Frame(self, bg="#0A1520", pady=5, padx=16)
        bar.pack(fill="x")

        tk.Label(bar, text="● GESTOS ATIVOS",
                 font=("Segoe UI", 7, "bold"),
                 bg="#0A1520", fg=SUCCESS).pack(side="left")

        self._lbl_gesto = tk.Label(
            bar,
            text="Aguardando gesto na câmera...",
            font=("Segoe UI", 7),
            bg="#0A1520", fg=TEXT_DIM,
        )
        self._lbl_gesto.pack(side="left", padx=(12, 0))

        tk.Label(bar,
                 text="PINÇA = selecionar arquivos  |  OK = executar macro",
                 font=("Segoe UI", 7),
                 bg="#0A1520", fg=TEXT_MUT).pack(side="right")

    def _set_gesto_status(self, msg, cor=None):
        self._lbl_gesto.config(text=msg, fg=cor or TEXT_DIM)

    # ── CALLBACKS DE GESTO ──────────────────

    def gesto_selecionar(self):
        """Pinça detectada: abre diálogos de seleção para a aba ativa."""
        self._set_gesto_status("PINÇA detectada — selecionando arquivos...", ACCENT)
        if self._active_tab == "comparar":
            self._selecionar_comparar()
        else:
            self._selecionar_converter()

    def gesto_executar(self):
        """OK detectado: executa o macro da aba ativa."""
        self._set_gesto_status("OK detectado — executando...", SUCCESS)
        if self._active_tab == "comparar":
            self._run_comparar()
        else:
            self._run_converter()

    def _selecionar_comparar(self):
        p = filedialog.askopenfilename(
            title="Planilha NOVA (mais recente)",
            filetypes=[("CSV / Excel", "*.csv *.xlsx *.xls"), ("Todos", "*.*")])
        if p:
            self.v_novo.set(p)
        else:
            self._set_gesto_status("Seleção cancelada.", TEXT_DIM)
            return

        p = filedialog.askopenfilename(
            title="Planilha ANTIGA (referência)",
            filetypes=[("CSV / Excel", "*.csv *.xlsx *.xls"), ("Todos", "*.*")])
        if p:
            self.v_antigo.set(p)
        else:
            self._set_gesto_status("Seleção cancelada.", TEXT_DIM)
            return

        p = filedialog.asksaveasfilename(
            title="Salvar resultado em",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if p:
            self.v_out_c.set(p)

        self._set_gesto_status("Arquivos selecionados. Faça OK para executar.", ACCENT)

    def _selecionar_converter(self):
        p = filedialog.askopenfilename(
            title="Arquivo CSV de entrada",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if p:
            self.v_csv.set(p)
        else:
            self._set_gesto_status("Seleção cancelada.", TEXT_DIM)
            return

        p = filedialog.asksaveasfilename(
            title="Salvar Excel em",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if p:
            self.v_out_v.set(p)

        self._set_gesto_status("Arquivo selecionado. Faça OK para executar.", ACCENT)

    # ── HEADER ──────────────────────────────

    def _build_header(self):
        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")

        h = tk.Frame(self, bg=SURFACE, pady=14, padx=24)
        h.pack(fill="x")

        # Ícone
        tk.Label(h, text="ESTOQUE", font=("Segoe UI", 7, "bold"),
                 bg=ACCENT, fg=BTN_FG, padx=8, pady=5).pack(side="left", padx=(0, 14))

        blk = tk.Frame(h, bg=SURFACE)
        blk.pack(side="left")
        tk.Label(blk, text="Gestao de Estoque de Calcados",
                 font=("Segoe UI", 13, "bold"), bg=SURFACE, fg=TEXT).pack(anchor="w")
        tk.Label(blk, text="Comparacao de planilhas  |  Conversor CSV para Excel",
                 font=("Segoe UI", 8), bg=SURFACE, fg=TEXT_DIM).pack(anchor="w")

        tk.Label(h, text=" v2.0 ", font=("Segoe UI", 8, "bold"),
                 bg=ACCENT, fg=BTN_FG, padx=6, pady=3).pack(side="right")

    # ── TABS ────────────────────────────────

    def _build_tabs(self):
        self._tab_btns   = {}
        self._tab_frames = {}

        # Barra
        bar = tk.Frame(self, bg=SURFACE, padx=16)
        bar.pack(fill="x")
        tk.Frame(bar, bg=BORDER, height=1).pack(fill="x", side="bottom")

        tabs = [
            ("comparar",  "  Comparar Estoque  "),
            ("converter", "  CSV para Excel  "),
        ]
        for key, label in tabs:
            btn = tk.Label(bar, text=label, font=("Segoe UI", 9, "bold"),
                           padx=6, pady=10, cursor="hand2", bg=SURFACE)
            btn.pack(side="left")
            btn.bind("<Button-1>", lambda e, k=key: self._show_tab(k))
            self._tab_btns[key] = btn

        # Conteúdo
        self._content = tk.Frame(self, bg=BG, padx=24, pady=10)
        self._content.pack(fill="both", expand=True)

        f_comp = tk.Frame(self._content, bg=BG)
        f_conv = tk.Frame(self._content, bg=BG)
        self._tab_frames = {"comparar": f_comp, "converter": f_conv}

        self._build_comparar(f_comp)
        self._build_converter(f_conv)
        self._show_tab("comparar")

        # Footer
        foot = tk.Frame(self, bg=SURFACE, pady=7)
        foot.pack(fill="x", side="bottom")
        tk.Label(foot, text="Selecione os arquivos e clique no botao para processar",
                 font=("Segoe UI", 8), bg=SURFACE, fg=TEXT_MUT).pack()

    def _show_tab(self, key):
        self._active_tab = key
        for k, f in self._tab_frames.items():
            f.pack_forget()
        self._tab_frames[key].pack(fill="both", expand=True)

        for k, btn in self._tab_btns.items():
            if k == key:
                btn.config(bg=ACCENT, fg=BTN_FG)
            else:
                btn.config(bg=SURFACE, fg=TEXT_DIM)

    # ── ABA COMPARAR ────────────────────────

    def _build_comparar(self, f):
        self.v_novo   = tk.StringVar()
        self.v_antigo = tk.StringVar()
        self.v_out_c  = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "Desktop", "comparacao_estoque.xlsx"))

        FileRow(f, "PLANILHA NOVA  (mais recente)", self.v_novo).pack(fill="x")
        FileRow(f, "PLANILHA ANTIGA  (referencia)", self.v_antigo).pack(fill="x")
        FileRow(f, "SALVAR RESULTADO EM", self.v_out_c, save_mode=True).pack(fill="x")

        tk.Frame(f, bg=BORDER, height=1).pack(fill="x", pady=(16, 8))

        self.log_c = LogBox(f, lines=6)
        self.log_c.pack(fill="x")
        self.log_c.log("Aguardando arquivos...", "dim")

        btn_frame = tk.Frame(f, bg=BG, pady=14)
        btn_frame.pack()
        self.btn_c = make_button(btn_frame, "  Comparar Planilhas  ",
                                 self._run_comparar, primary=True, width=28)
        self.btn_c.pack(ipady=4)

    # ── ABA CONVERTER ───────────────────────

    def _build_converter(self, f):
        self.v_csv   = tk.StringVar()
        self.v_out_v = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "Desktop", "planilha_convertida.xlsx"))

        FileRow(f, "ARQUIVO CSV DE ENTRADA", self.v_csv).pack(fill="x")
        FileRow(f, "SALVAR EXCEL EM", self.v_out_v, save_mode=True).pack(fill="x")

        tk.Frame(f, bg=BORDER, height=1).pack(fill="x", pady=(16, 8))

        # Card informativo
        card = tk.Frame(f, bg=SURFACE2, padx=14, pady=10,
                        highlightbackground=BORDER, highlightthickness=1)
        card.pack(fill="x", pady=(0, 10))
        tk.Label(card, text="O que este conversor faz:",
                 font=("Segoe UI", 9, "bold"), bg=SURFACE2, fg=ACCENT).pack(anchor="w")
        for txt in [
            "  - Detecta encoding automatico (Windows-1252, UTF-8, Latin-1)",
            "  - Converte separador ; para colunas do Excel",
            "  - Aplica cabecalho formatado e filtros automaticos",
            "  - Ajusta largura das colunas automaticamente",
        ]:
            tk.Label(card, text=txt, font=("Segoe UI", 8),
                     bg=SURFACE2, fg=TEXT_DIM).pack(anchor="w", pady=1)

        self.log_v = LogBox(f, lines=4)
        self.log_v.pack(fill="x")
        self.log_v.log("Aguardando arquivo CSV...", "dim")

        btn_frame = tk.Frame(f, bg=BG, pady=14)
        btn_frame.pack()
        self.btn_v = make_button(btn_frame, "  Converter CSV para Excel  ",
                                 self._run_converter, primary=True, width=30)
        self.btn_v.pack(ipady=4)

    # ── AÇÕES ───────────────────────────────

    def _run_comparar(self):
        novo   = self.v_novo.get().strip()
        antigo = self.v_antigo.get().strip()
        saida  = self.v_out_c.get().strip()
        if not novo or not antigo:
            messagebox.showwarning("Atencao", "Selecione as duas planilhas.")
            return

        self.log_c.clear()
        self.log_c.log("Lendo e processando planilhas...", "info")
        self.btn_c.config(text="  Processando...  ", state="disabled")
        self.update()

        def run():
            try:
                comparar_planilhas(novo, antigo, saida)
                self.after(0, lambda: self._done_c(saida, None))
            except Exception as e:
                self.after(0, lambda: self._done_c(saida, str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _done_c(self, saida, err):
        self.btn_c.config(text="  Comparar Planilhas  ", state="normal")
        if err:
            self.log_c.log(f"ERRO: {err}", "err")
            self._set_gesto_status("Erro na execução.", DANGER)
            messagebox.showerror("Erro", err)
        else:
            self.log_c.log("Comparacao concluida com sucesso!", "ok")
            self.log_c.log(f"Arquivo: {saida}", "dim")
            self._set_gesto_status("Concluído com sucesso!", SUCCESS)
            messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{saida}")

    def _run_converter(self):
        csv   = self.v_csv.get().strip()
        saida = self.v_out_v.get().strip()
        if not csv:
            messagebox.showwarning("Atencao", "Selecione um arquivo CSV.")
            return

        self.log_v.clear()
        self.log_v.log("Convertendo arquivo...", "info")
        self.btn_v.config(text="  Processando...  ", state="disabled")
        self.update()

        def run():
            try:
                csv_para_excel(csv, saida)
                self.after(0, lambda: self._done_v(saida, None))
            except Exception as e:
                self.after(0, lambda: self._done_v(saida, str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _done_v(self, saida, err):
        self.btn_v.config(text="  Converter CSV para Excel  ", state="normal")
        if err:
            self.log_v.log(f"ERRO: {err}", "err")
            self._set_gesto_status("Erro na execução.", DANGER)
            messagebox.showerror("Erro", err)
        else:
            self.log_v.log("Conversao concluida com sucesso!", "ok")
            self.log_v.log(f"Arquivo: {saida}", "dim")
            self._set_gesto_status("Concluído com sucesso!", SUCCESS)
            messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{saida}")


if __name__ == "__main__":
    app = App()
    app.mainloop()